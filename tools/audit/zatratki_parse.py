import openpyxl, glob, os, re, json
from collections import defaultdict
ROOT="/Users/jakedaurenbekov/Projects/myata-finance/docs/Accounting/1. ОПиУ и ДДС/Затратки/"
files=sorted(glob.glob(ROOT+"20*/*.xlsx"))
NAME_RE=re.compile(r"^\s*(\d{1,2})[.,]+(\d{1,2})\.?\s*$")
def norm(s): return re.sub(r"\s+"," ",str(s)).strip().rstrip(':').strip().lower()
res={}; bad=[]; skipped=[]
for f in files:
    year=int(f.split("/")[-2]); fm=int(os.path.basename(f)[:2]); key=f"{year}-{fm:02d}"
    if key=="2022-07": continue   # старый формат (июль 2022) — берём из PnL_Daily
    wb=openpyxl.load_workbook(f, read_only=True, data_only=True)
    dated=[ws for ws in wb.worksheets if NAME_RE.match(ws.title)]
    days_in_month={int(NAME_RE.match(ws.title).group(1)) for ws in dated if int(NAME_RE.match(ws.title).group(2))==fm}
    sheets=[]
    for pos,ws in enumerate(dated):
        d,m=map(int, NAME_RE.match(ws.title).groups())
        if m==fm: sheets.append((d,ws))
        elif d not in days_in_month and abs(pos-(d-1))<=2: sheets.append((d,ws))
        else: skipped.append((key,ws.title))
    days=[]
    for d,ws in sheets:
        grid=[row for row in ws.iter_rows(min_row=1,max_row=20,min_col=1,max_col=14,values_only=True)]
        lab={}
        for r,row in enumerate(grid):
            for c in range(4,9):
                v=row[c] if c<len(row) else None
                if isinstance(v,str) and norm(v): lab.setdefault(norm(v),(r,c))
        def num(r,c):
            v=grid[r][c] if r<len(grid) and c<len(grid[r]) else None
            return float(v) if isinstance(v,(int,float)) else None
        def right(label):
            p=lab.get(label)
            if not p: return None
            r,c=p
            for cc in range(c+1,c+4):
                v=num(r,cc)
                if v is not None: return v
            return None
        total=right("итого"); start=right("на начало дня")
        exp=right("затраты"); inflow=right("приход"); coll=right("инкассация"); refund=right("возврат"); left=right("остаток")
        bank=cashless=cash=None
        hp=lab.get("банк")
        if hp:
            r,c=hp
            for cc in range(c,c+4):
                h=grid[r][cc] if cc<len(grid[r]) else None
                if not isinstance(h,str): continue
                hn=norm(h); v=num(r+1,cc)
                if hn=="банк": bank=v
                elif hn in ("без нал","безнал"): cashless=v
                elif hn=="нал": cash=v
        end=left
        if end is None and "возврат" in lab and "имя" in lab:
            r1=lab["возврат"][0]; r2=lab["имя"][0]
            for r in range(r1+1,r2):
                v=num(r,6)
                if v is not None: end=v; break
        if total is None and bank is None: bad.append((key,ws.title)); continue
        ch={}
        grid2=[row for row in ws.iter_rows(min_row=1,max_row=45,min_col=1,max_col=16,values_only=True)]
        for r in (0,1):
            for c in range(7,14):
                h=grid2[r][c] if c<len(grid2[r]) else None
                if not isinstance(h,str): continue
                hn=norm(h)
                name=None
                if "касп" in hn: name="kaspi"
                elif "халык" in hn or "народн" in hn or "halyk" in hn: name="halyk"
                elif "вольт" in hn or "wolt" in hn: name="wolt"
                elif "налич" in hn: name="cash_hdr"
                elif "глово" in hn or "яндекс" in hn or "перевод" in hn: name="other_ch"
                if not name: continue
                sm=0.0
                for rr in range(r+1,len(grid2)):
                    v=grid2[rr][c] if c<len(grid2[rr]) else None
                    if isinstance(v,str) and v.strip(): break
                    if isinstance(v,(int,float)): sm+=float(v)
                ch[name]=ch.get(name,0)+sm
        days.append(dict(kaspi=ch.get("kaspi",0),halyk=ch.get("halyk",0),wolt=ch.get("wolt",0),cash_hdr=ch.get("cash_hdr",0),other_ch=ch.get("other_ch",0),d=d,total=total or 0,bank=bank or 0,cashless=cashless or 0,cash=cash or 0,exp=exp or 0,inflow=inflow or 0,coll=coll or 0,refund=refund or 0,start=start,end=end))
    days.sort(key=lambda x:x['d'])
    a=defaultdict(float)
    for x in days:
        for k in ("total","bank","cashless","cash","exp","inflow","coll","refund","kaspi","halyk","wolt","cash_hdr","other_ch"): a[k]+=x[k]
    a["days"]=len(days); a["zero_days"]=sum(1 for x in days if x["total"]==0)
    both=[x for x in days if x["start"] is not None and x["end"] is not None]
    a["loop_days"]=len(both)
    a["d_end_start"]=sum(x["end"]-x["start"] for x in both)
    a["implied"]=sum(x["cash"]-x["exp"]+x["inflow"]-x["coll"]+x["refund"] for x in both)
    drift=0; nd=0
    for p,q in zip(days,days[1:]):
        if p["end"] is not None and q["start"] is not None:
            nd+=1; drift+= q["start"]-p["end"]
    a["drift"]=drift; a["drift_pairs"]=nd
    a["end_last"]=next((x["end"] for x in reversed(days) if x["end"] is not None), None)
    a["start_first"]=next((x["start"] for x in days if x["start"] is not None), None)
    res[key]={"agg":dict(a),"days":days}
json.dump(res, open("zatratki_monthly.json","w"), ensure_ascii=False)
print(f"{'month':8}{'days':>5}{'zero':>5}{'total':>13}{'bank':>13}{'cashless':>10}{'cash':>12}{'expenses':>12}{'inflow':>10}{'collect':>10}{'refund':>8}{'loopD':>6}{'Σend-start':>12}{'implied':>12}{'unexpl':>12}{'drift':>11}{'endLast':>11}")
for k in sorted(res):
    a=res[k]["agg"]; g=lambda n: round(a.get(n) or 0)
    print(f"{k:8}{g('days'):>5}{g('zero_days'):>5}{g('total'):>13,}{g('bank'):>13,}{g('cashless'):>10,}{g('cash'):>12,}{g('exp'):>12,}{g('inflow'):>10,}{g('coll'):>10,}{g('refund'):>8,}{g('loop_days'):>6}{g('d_end_start'):>12,}{g('implied'):>12,}{g('d_end_start')-g('implied'):>12,}{g('drift'):>11,}{g('end_last'):>11,}")
print("unparsed:", len(bad), bad[:12]); print("skipped (other month):", len(skipped), skipped[:40])
