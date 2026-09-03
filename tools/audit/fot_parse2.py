import openpyxl, re, glob, os, sys
from collections import defaultdict
ROOT="/Users/jakedaurenbekov/Projects/myata-finance/docs/Accounting/ФОТ/2026/"
def num(v):
    if isinstance(v,(int,float)): return float(v)
    if isinstance(v,str):
        try: return float(v.replace(" ","").replace(",","."))
        except: return 0.0
    return 0.0
def dept_of(pos):
    p=str(pos or "").strip().lower()
    if re.fullmatch(r"0[.,]\d+", p): return "Зал"
    if any(k in p for k in ["управля","бухгалт","калькул","главбух","глав.бух"]): return "Менеджмент"
    if "смм" in p: return "СММ"
    if "развоз" in p or "водител" in p: return "Развозка"
    if "админ" in p or "менеджер" in p or p in ("оф","официант"): return "Зал"
    if "кальян" in p or "кальн" in p: return "Кальян"
    if "бар" in p: return "Бар"
    if "повар" in p or "су шеф" in p or "шеф" in p: return "Кухня"
    if "тех" in p: return "Прочее"
    return "?"+p
def pick(wb, fname):
    m=re.search(r"(\d\d)\.(\d\d)\.20\d\d", fname); day,mm=int(m.group(1)),m.group(2)
    for ws in wb.worksheets:
        t=ws.title.replace(" ","")
        if day<=15 and re.match(rf"^0?1\.{mm}(\.\d\d)?-15\.{mm}", t): return ws.title
        if day>15 and re.match(rf"^16\.{mm}-", t): return ws.title
def parse(wsf, wsv):
    res=defaultdict(lambda:[0.0,0.0,0]); people=[]; headers=0; computed=0
    for r in range(3, 60):
        name=wsv.cell(r,2).value; pos=wsv.cell(r,3).value
        if isinstance(name,str) and name.strip().upper().startswith("ФИО"):
            headers+=1
            if headers>=2: break
            continue
        if not name or not str(name).strip() or str(name).startswith("ТОТАЛ"): continue
        w=wsv.cell(r,23).value
        U=num(wsv.cell(r,21).value); V=num(wsv.cell(r,22).value)
        T=sum(num(wsv.cell(r,c).value) for c in range(4,20))
        if isinstance(w,(int,float)) and 0<=w<5_000_000: acc=float(w)
        elif U or V: acc=T*U+V; computed+=1          # «начислено» не заполнено/не закэшировано — считаем по формуле смены×оклад+%
        else: acc=0.0
        if acc>=5_000_000: print(f"   ⚠ мусор в строке {r} ({name}): {acc:,.0f} — обнулено"); acc=0.0
        if isinstance(w,(int,float)) and w>=5_000_000: print(f"   ⚠ мусор в W{r} ({name}): {w} — пересчитано как {acc:,.0f}")
        adv=num(wsv.cell(r,24).value)
        if acc==0 and adv==0: continue
        d=dept_of(pos); res[d][0]+=acc; res[d][1]+=adv; res[d][2]+=1
        people.append((str(name).strip(), str(pos or "").strip(), d, acc, adv))
    return res, people, computed
def all_periods():
    files=sorted(f for f in glob.glob(ROOT+"*/**/*.xlsx", recursive=True) if os.path.basename(f).startswith("ЗП"))
    best={}   # sheet title -> (total, file, res, people)
    for f in files:
        fn=os.path.basename(f)
        wbf=openpyxl.load_workbook(f); wbv=openpyxl.load_workbook(f, data_only=True)
        for ws in wbv.worksheets:
            t=ws.title.replace(" ","")
            m=re.match(r"^0?(\d{1,2})\.(\d{2})", t)
            if not m: continue
            res,people,computed=parse(wbf[ws.title], ws)
            tot=sum(v[0] for v in res.values())
            own = pick(wbv, fn)==ws.title if re.search(r"\d\d\.\d\d\.20\d\d", fn) else False
            cur=best.get(t)
            if cur is None or (own and tot>=1_500_000) or (not cur[4] and tot>cur[0]): best[t]=(tot,fn,res,people,own and tot>=1_500_000)
    return best
if __name__=="__main__" and "--all" in sys.argv:
    best=all_periods(); grand=defaultdict(lambda: defaultdict(float)); mg=defaultdict(list)
    for t,(tot,fn,res,people,own) in sorted(best.items(), key=lambda x: (x[0].split("-")[0].split(".")[1], x[0])):
        mm=re.match(r"^0?\d{1,2}\.(\d{2})", t).group(1); key=("2025-" if mm=="12" else "2026-")+mm
        for d,(a,ad,n) in res.items(): grand[key][d]+=a
        mg[key]+=[(p[0],round(p[3])) for p in people if p[2]=="Менеджмент"]
        print(f"{t:16} из {fn:30} начислено {tot:>11,.0f} чел {sum(v[2] for v in res.values()):2}  без должности: {[(p[0],round(p[3])) for p in people if p[2].startswith('?')]}")
    depts=["Менеджмент","Зал","Бар","Кухня","Кальян","Развозка","СММ","Прочее"]
    print(f"\n{'месяц':8}"+"".join(f"{d:>12}" for d in depts)+f"{'ИТОГО':>12}")
    for k in sorted(grand): print(f"{k:8}"+"".join(f"{grand[k].get(d,0):>12,.0f}" for d in depts)+f"{sum(grand[k].values()):>12,.0f}")
    unk=set(d for k in grand for d in grand[k] if d.startswith("?"))
    if unk: print("⚠ нераспознанные должности:", unk)
    sys.exit(0)
if __name__=="__main__":
    files=sorted(f for f in glob.glob(ROOT+"*/**/*.xlsx", recursive=True) if os.path.basename(f).startswith("ЗП"))
    grand=defaultdict(lambda: defaultdict(float)); mg=defaultdict(list)
    for f in files:
        fn=os.path.basename(f); t=pick(openpyxl.load_workbook(f, read_only=True), fn)
        if not t: print("no sheet:", fn); continue
        wsf=openpyxl.load_workbook(f)[t]; wsv=openpyxl.load_workbook(f, data_only=True)[t]
        res,people,computed=parse(wsf,wsv)
        m=re.search(r"\d\d\.(\d\d)\.20\d\d", fn); key="2026-"+m.group(1)
        for d,(a,ad,n) in res.items(): grand[key][d]+=a
        mg[key]+=[(p[0],round(p[3])) for p in people if p[2]=="Менеджмент"]
        print(f"{fn:30} {t:14} начислено {sum(v[0] for v in res.values()):>11,.0f} чел {sum(v[2] for v in res.values()):2} (вычислено по формуле: {computed})")
    depts=["Менеджмент","Зал","Бар","Кухня","Кальян","Развозка","СММ","Прочее"]
    print(f"\n{'месяц':8}"+"".join(f"{d:>12}" for d in depts)+f"{'ИТОГО':>12}")
    for k in sorted(grand): print(f"{k:8}"+"".join(f"{grand[k].get(d,0):>12,.0f}" for d in depts)+f"{sum(grand[k].values()):>12,.0f}")
    for k in sorted(mg): print(k, "менеджмент:", mg[k])
    unk=[d for k in grand for d in grand[k] if d.startswith("?")]
    if unk: print("⚠ нераспознанные должности:", set(unk))
