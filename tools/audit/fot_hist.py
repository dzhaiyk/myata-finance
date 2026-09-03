import openpyxl, re, glob, os, json, sys
from collections import defaultdict
ROOT="/Users/jakedaurenbekov/Projects/myata-finance/docs/Accounting/ФОТ/"
MRE=re.compile(r"^\s*(январ|феврал|март|апрел|ма[йя]|июн|июл|август|сентябр|октябр|ноябр|декабр)[а-яё]*\s*с?\s*(\d{1,2})\s*-\s*(\d{1,2})", re.I)
MN={'январ':1,'феврал':2,'март':3,'апрел':4,'май':5,'мая':5,'июн':6,'июл':7,'август':8,'сентябр':9,'октябр':10,'ноябр':11,'декабр':12}
def num(v):
    if isinstance(v,(int,float)): return float(v)
    if isinstance(v,str):
        t=v.replace(" ","").replace(",",".")
        if not re.fullmatch(r"-?\d+(\.\d+)?", t): return 0.0
        return float(t)
    return 0.0
def dept_of(name,pos):
    p=str(pos if pos is not None else "").strip().lower(); n=str(name or "").strip().lower()
    if isinstance(pos,(int,float)) or re.fullmatch(r"0[.,]\d+", p) or p=="0": return "Зал"
    if "водител" in n or "развоз" in n or "развоз" in p or "водител" in p: return "Развозка"
    if any(k in p for k in ["управля","бухгалт","калькул","главбух","глав.бух","гл бух","гл. бух"]): return "Менеджмент"
    if "смм" in p or "смм" in n: return "СММ"
    if "админ" in p or "менеджер" in p or p in ("оф","официант","офф") or "оф" in p and "%" in p: return "Зал"
    if "кальян" in p or "кальн" in p: return "Кальян"
    if "бар" in p: return "Бар"
    if "повар" in p or "су шеф" in p or "шеф" in p: return "Кухня"
    if "тех" in p: return "Прочее"
    if "раннер" in p or "хостес" in p: return "Зал"
    if p in ("оф.",): return "Зал"
    return "?"+p
def file_ym(path):
    rel=path[len(ROOT):]; fy=int(rel.split("/")[0]); fn=os.path.basename(path)
    m=re.search(r"(\d\d)\.(\d\d)\.(20\d\d)", fn)
    if m: return int(m.group(3)), int(m.group(2))
    m=re.match(r"(\d\d)\.", fn)
    if m: return fy, int(m.group(1))
    d=rel.split("/")[1] if "/" in rel else ""
    m=re.match(r"(\d\d)\.", d)
    return fy, (int(m.group(1)) if m else 1)
def sheet_period(title, fy, fm):
    m=MRE.match(title.replace("ё","е"))
    if not m: return None
    mon=MN[m.group(1).lower()] if m.group(1).lower() in MN else MN.get(m.group(1).lower()[:5])
    if mon is None:
        for k,v in MN.items():
            if m.group(1).lower().startswith(k[:3]): mon=v; break
    d1,d2=int(m.group(2)),int(m.group(3))
    if mon==fm+1: best=fy
    elif fm==12 and mon==1: best=fy+1
    else: best=fy if mon<=fm else fy-1
    return best, mon, (1 if d2<=16 else 2), d1, d2
def parse_sheet(ws):
    rows=list(ws.iter_rows(min_row=1, max_row=min(ws.max_row,120), max_col=45, values_only=True))
    res=defaultdict(float); people=[]; block=None; prev=None; sheet_total=0.0; blocks=0; diag=[]
    for r,row in enumerate(rows):
        name=row[1] if len(row)>1 else None; pos=row[2] if len(row)>2 else None
        sname=str(name).strip() if name is not None else ""
        if sname.lower() in ("фио","имя"):
            diag.append((r+1,[(c,str(v)[:12]) for c,v in enumerate(row) if isinstance(v,str) and v.strip() and c>=15]))
            labels={c:str(v).strip().lower() for c,v in enumerate(row) if isinstance(v,str) and v.strip()}
            if r>0 and not any(c>=15 and not re.fullmatch(r"(пн|вт|ср|чт|пт|сб|вс)", l) for c,l in labels.items()):
                for c,v in enumerate(rows[r-1]):
                    if isinstance(v,str) and v.strip() and c>=15: labels.setdefault(c,str(v).strip().lower())
            if any(l.strip()=="тотал" for c,l in labels.items() if c>=15): block=None; break
            sh=next((c for c,l in labels.items() if c>=15 and l.startswith("смен")), None)
            acc=next((c for c,l in labels.items() if c>=15 and ("общ" in l or l.startswith("тотал зп"))), None)
            if acc is None: acc=next((c for c,l in labels.items() if c>=15 and (l.startswith("итого") or l.startswith("начисл"))), None)
            rate=next((c for c,l in labels.items() if c>=15 and (l.startswith("оклад") or l.startswith("за смену"))), None)
            pct=next((c for c,l in labels.items() if c>=15 and l.strip()=="%"), None)
            adv=next((c for c,l in labels.items() if c>=15 and l.startswith("аванс")), None)
            if sh is None and acc is None: block=None; continue
            if acc is None and rate is None and pct is None:
                if prev: block=dict(sh=sh if sh is not None else prev["sh"],acc=prev["acc"],rate=prev["rate"],pct=prev["pct"],adv=prev["adv"])
                elif sh is not None: block=dict(sh=sh,rate=sh+1,pct=sh+2,acc=sh+3,adv=sh+4)   # без подписей: смены|оклад|%|итого|авансы
                else: block=None; continue
            else: block=dict(sh=sh,acc=acc,rate=rate,pct=pct,adv=adv)
            prev=block; blocks+=1; continue
        if block is None: continue
        if not sname or sname.startswith("ТОТАЛ") or sname in (",","."," "):
            if block["acc"] is not None and len(row)>block["acc"] and isinstance(row[block["acc"]],(int,float)) and row[block["acc"]]>sheet_total: sheet_total=float(row[block["acc"]])
            continue
        if isinstance(name,(int,float)) and not isinstance(pos,str): continue
        w=row[block["acc"]] if block["acc"] is not None and len(row)>block["acc"] else None
        U=num(row[block["rate"]]) if block["rate"] is not None and len(row)>block["rate"] else 0.0
        V=num(row[block["pct"]]) if block["pct"] is not None and len(row)>block["pct"] else 0.0
        T=num(row[block["sh"]]) if block["sh"] is not None and len(row)>block["sh"] else 0.0
        if isinstance(w,(int,float)) and 0<=w<5_000_000: a=float(w)
        elif U or V: a=T*U+V
        else: a=0.0
        if a>=5_000_000 or a<0: a=0.0
        ad=num(row[block["adv"]]) if block["adv"] is not None and len(row)>block["adv"] else 0.0
        if a==0 and ad==0: continue
        d=dept_of(sname,pos); res[d]+=a; people.append((sname,str(pos),d,round(a)))
    return dict(res), people, sheet_total, blocks, diag
def parse_svod(ws):
    rows=list(ws.iter_rows(min_row=1, max_row=min(ws.max_row,80), max_col=13, values_only=True))
    per=None; res=defaultdict(float); people=[]
    for row in rows:
        a=row[0]
        if isinstance(a,str):
            m=re.search(r"с (\d\d)\.(\d\d)\.(20\d\d) по (\d\d)\.(\d\d)\.(20\d\d)", a)
            if m: per=(int(m.group(3)), int(m.group(2)), 1 if int(m.group(1))<=15 else 2, int(m.group(1)), int(m.group(4)))
    hdr=next((i for i,row in enumerate(rows) if row[0]=="Сотрудник"), None)
    if hdr is None or per is None: return None
    labels=[str(v).strip().lower() if v else "" for v in rows[hdr]]
    ci=labels.index("итого") if "итого" in labels else 9
    for row in rows[hdr+1:]:
        nm=row[0]; pos=row[1]; v=row[ci] if ci<len(row) else None
        if not isinstance(nm,str) or not isinstance(v,(int,float)) or v==0: continue
        d=dept_of(nm,pos); res[d]+=float(v); people.append((nm,str(pos),d,round(v)))
    return per, dict(res), people
files=sorted(f for f in glob.glob(ROOT+"20*/**/*.xlsx", recursive=True) if int(f[len(ROOT):][:4])<=2025 and re.search(r"Зарплата_|ЗП Мята", os.path.basename(f)))
periods={}  # (y,m,half) -> list of (total, file, sheet, res, people, sheet_total)
for f in files:
    fy,fm=file_ym(f)
    try: wb=openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as e: print("ERR",f,e); continue
    for ws in wb.worksheets:
        if ws.title.startswith("Сводный"):
            sv=parse_svod(ws)
            if sv:
                per,res,people=sv; tot=sum(res.values())
                periods.setdefault(per[:3],[]).append((tot, os.path.basename(f), ws.title+" (iiko)", res, people, tot, 1, True))
            continue
        if fy==2022 and ws.title=="Лист1":
            m=re.search(r"_(\d)\.xlsx$", os.path.basename(f)); half=int(m.group(1)) if m else 1
            per=(fy,fm,half,1 if half==1 else 16,15 if half==1 else 31)
        else:
            p=sheet_period(ws.title, fy, fm)
            if not p: continue
            per=p
            m2=re.search(r"_(\d)\.xlsx$", os.path.basename(f))
            if m2 and fy==2023 and len(wb.sheetnames)<=4 and p[2]!=int(m2.group(1)) and p[1]==fm: per=(p[0],p[1],int(m2.group(1)),p[3],p[4])
        res,people,st,blocks,diag=parse_sheet(ws)
        tot=sum(res.values())
        if tot==0:
            print(f"  ∅ {os.path.basename(f)} [{ws.title}] period={per[:3]} blocks={blocks} headers={diag[:3]}"); continue
        fnb=os.path.basename(f); md=re.search(r"(\d\d)\.\d\d\.20\d\d", fnb)
        fhalf = (1 if int(md.group(1))<=16 else 2) if md else (1 if fnb.endswith("_1.xlsx") else 2 if fnb.endswith("_2.xlsx") else None)
        own = (per[1]==fm and per[0]==fy and fhalf==per[2])
        periods.setdefault(per[:3],[]).append((tot, os.path.basename(f), ws.title, res, people, st, blocks, own))
pe=json.load(open("pnl_excel.json"))
chosen={}
print("выбор ведомости по периодам (при нескольких копиях — максимальная сумма):")
for k in sorted(periods):
    cands=sorted(periods[k], key=lambda x:(-(x[7] and 1_500_000<=x[0]<12_000_000), -(x[0]<12_000_000), -x[0])); best=cands[0]; chosen[k]=best
    alt="" if len(cands)==1 else " | другие копии: "+", ".join(f"{c[1]}[{c[2]}]={c[0]:,.0f}" for c in cands[1:] if abs(c[0]-best[0])>0.02*best[0])
    unk=[p for p in best[4] if p[2].startswith("?")]
    print(f"  {k[0]}-{k[1]:02d} h{k[2]}: {best[0]:>11,.0f} (в листе итог {best[5]:>11,.0f}, блоков {best[6]}) из {best[1]} [{best[2]}]{alt}" + (f"  ⚠ без отдела: {unk}" if unk else ""))
DEPTS=["Менеджмент","Зал","Бар","Кухня","Кальян","Развозка","СММ","Прочее"]
month=defaultdict(lambda: defaultdict(float)); halves=defaultdict(set)
for (y,m,h),best in chosen.items():
    for d,v in best[3].items(): month[(y,m)][d]+=v
    halves[(y,m)].add(h)
print(f"\n{'месяц':8}{'половины':>9}"+"".join(f"{d:>11}" for d in DEPTS)+f"{'ВЕДОМОСТИ':>12}{'Excel ФОТ':>12}{'разница':>12}")
out={}
for (y,m) in sorted(month):
    tot=sum(month[(y,m)].values()); ex=pe.get(str(y),{}).get("ФОТ",[0]*12)[m-1] or 0
    hs="".join(str(h) for h in sorted(halves[(y,m)]))
    out[f"{y}-{m:02d}"]=dict(halves=hs, total=tot, excel=ex, **{d:month[(y,m)].get(d,0) for d in DEPTS})
    print(f"{y}-{m:02d}{hs:>9}"+"".join(f"{month[(y,m)].get(d,0):>11,.0f}" for d in DEPTS)+f"{tot:>12,.0f}{ex:>12,.0f}{ex-tot:>12,.0f}")
json.dump(out, open("fot_hist.json","w"), ensure_ascii=False)
