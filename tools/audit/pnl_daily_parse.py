import openpyxl, json, re, sys
from collections import defaultdict, OrderedDict
ROOT="/Users/jakedaurenbekov/Projects/myata-finance/docs/Accounting/1. ОПиУ и ДДС/PnL daily/"
MONTHS={'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12,
        'янв':1,'фев':2,'мар':3,'апр':4,'май':5,'мая':5,'июн':6,'июл':7,'авг':8,'сен':9,'окт':10,'ноя':11,'дек':12}
def month_of(title):
    t=title.strip().lower()
    m=re.match(r'([a-zа-я]{3})', t)
    return MONTHS.get(m.group(1)) if m else None
SKIP={'день','объёмы деятельности','объемы деятельности'}
out={}; daily={}
for y in (2022,2023,2024,2025):
    wb=openpyxl.load_workbook(f"{ROOT}PnL_Daily_{y}.xlsx", read_only=True, data_only=True)
    for ws in wb.worksheets:
        m=month_of(ws.title)
        if not m: continue
        rows=list(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=40, values_only=True))
        hdr=None
        for row in rows[:5]:
            for j,v in enumerate(row):
                if isinstance(v,str) and v.strip().lower().startswith('месяц'): hdr=j
        lines=OrderedDict(); rev=[]
        for row in rows:
            nm=None; ni=None
            for j in range(0,3):
                v=row[j] if j<len(row) else None
                if isinstance(v,str) and v.strip(): nm=v; ni=j; break
            if nm is None: continue
            name=re.sub(r'\s+',' ',nm).strip()
            if name.lower() in SKIP: continue
            vals=[v if isinstance(v,(int,float)) else 0 for v in row[ni+1:ni+32]]
            tot=row[hdr] if (hdr is not None and hdr<len(row) and isinstance(row[hdr],(int,float))) else sum(vals)
            key=name; i=2
            while key in lines: key=f"{name}#{i}"; i+=1
            lines[key]=round(tot)
            if name.startswith('ДОХОДЫ'): rev=[round(v) for v in vals]
        out.setdefault(y,{})[m]=lines; daily[(y,m)]=rev
json.dump({"monthly":{str(y):{str(m):v for m,v in d.items()} for y,d in out.items()}, "daily_revenue":{f"{k[0]}-{k[1]:02d}":v for k,v in daily.items()}}, open("pnl_daily.json","w"), ensure_ascii=False)
KEY=['ДОХОДЫ тенге','Кухня','Бар','Кальян','Прочее','РАСХОДЫ тенге','CapEx (инвестиции)','Паушальный взнос','Ремонт','Мебель и техника','CAPEX прочее','OpEx (ежемесячные расходы)','OpEx (ежедневные расходы)','ФОТ','ФОТ Прочее','Развозка','Taxes','Закуп','Food cost','Закуп кухня','Закуп бар','Закуп кальян','Хозтовары','Маркетинг','СММ','Таргет','Маркетинг прочее','Аренда','Коммуслуги','Коммунальные платежи','BI Service','Интернет и связь','Ком услуги прочее','Комиссий банка','Налоги','Роялти','Software','OpEx прочее','Меню','Дератизация/дезинсекция','Чистка жироуловителей','Мелкий ремонт','Форма для персонала','Прочее#2','Прочие расходы']
print(f"{'line':28}"+"".join(f"{y:>14}" for y in (2022,2023,2024,2025)))
for k in KEY:
    row=[]
    for y in (2022,2023,2024,2025):
        s=0
        for m,lines in out.get(y,{}).items():
            s+=lines.get(k,0)
        row.append(s)
    if any(row): print(f"{k:28}"+"".join(f"{v:>14,}" for v in row))
print("\nmonths present:", {y: sorted(out[y].keys()) for y in out})
print("\nmonthly: revenue / expenses (PnL_Daily)")
for y in (2022,2023,2024,2025):
    for m in sorted(out[y]):
        L=out[y][m]; rev=L.get('ДОХОДЫ тенге',0); exp=L.get('РАСХОДЫ тенге', L.get('OpEx (ежемесячные расходы)', L.get('OpEx (ежедневные расходы)',0)))
        print(f"{y}-{m:02d} rev {rev:>13,} exp {exp:>12,}  ФОТ {L.get('ФОТ',0):>11,}  закуп {L.get('Закуп',L.get('Food cost',0)):>11,}")
