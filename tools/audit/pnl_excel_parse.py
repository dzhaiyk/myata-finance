import openpyxl, json, re
ROOT="/Users/jakedaurenbekov/Projects/myata-finance/docs/Accounting/1. ОПиУ и ДДС/"
wb=openpyxl.load_workbook(ROOT+"1. PnL.xlsx", read_only=True, data_only=True)
out={}
for y in ("2022","2023","2024","2025"):
    ws=wb[f"PnL_{y}"]; lines={}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=15, values_only=True):
        name=row[0]
        if not isinstance(name,str) or not name.strip(): continue
        name=re.sub(r"\s+"," ",name).strip()
        key=name; i=2
        while key in lines: key=f"{name}#{i}"; i+=1
        lines[key]=[v if isinstance(v,(int,float)) else 0 for v in row[1:13]]
    out[y]=lines
json.dump(out, open("pnl_excel.json","w"), ensure_ascii=False)
for y,l in out.items(): print(y, len(l), "lines; revenue", round(sum(l["ДОХОДЫ тенге"])), "expenses", round(sum(l["РАСХОДЫ тенге"])))
